from flask import Flask, render_template, Response
from flask_socketio import SocketIO, emit
import cv2
import mediapipe as mp
import numpy as np

from flask import Flask, render_template, Response, request, redirect, url_for, session, flash
from flask_socketio import SocketIO, emit
import cv2
import mediapipe as mp
import numpy as np
import os
from datetime import datetime

app = Flask(__name__)
app.secret_key = 'your_secret_key'  # Needed for session management
socketio = SocketIO(app)

import openpyxl
from openpyxl import Workbook
import os

# In-memory user storage: username -> {email, password}
users = {}

# Initialize MediaPipe Hands
mp_hands = mp.solutions.hands
hands = mp_hands.Hands(max_num_hands=1)
mp_drawing = mp.solutions.drawing_utils

# Drawing variables
canvas = np.zeros((480, 640, 3), np.uint8)
xp, yp = 0, 0
drawing = False
current_color = (0, 0, 255)  # Default red
camera_active = True  # Camera state flag

@app.route('/')
def home():
    if 'username' in session:
        return redirect(url_for('painting'))
    else:
        return redirect(url_for('index'))

@app.route('/index')
def index():
    return render_template('index.html')

@app.route('/register', methods=['GET', 'POST'])
def register():
    excel_file = 'users.xlsx'

    if request.method == 'POST':
        username = request.form.get('username')
        email = request.form.get('email')
        password = request.form.get('password')
        confirm_password = request.form.get('confirm_password')

        if not username or not email or not password or not confirm_password:
            flash('Please fill out all fields.')
            return redirect(url_for('register'))

        if password != confirm_password:
            flash('Passwords do not match.')
            return redirect(url_for('register'))

        # Check if Excel file exists, create if not
        if not os.path.exists(excel_file):
            wb = Workbook()
            ws = wb.active
            ws.title = "Users"
            ws.append(['Username', 'Email', 'Password'])
            wb.save(excel_file)

        # Load workbook and check for existing username
        wb = openpyxl.load_workbook(excel_file)
        ws = wb['Users']

        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == username:
                flash('Username already exists.')
                wb.close()
                return redirect(url_for('register'))

        # Append new user data
        ws.append([username, email, password])
        wb.save(excel_file)
        wb.close()

        flash('Registration successful. Please log in.')
        return redirect(url_for('login'))

    return render_template('register.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    excel_file = 'users.xlsx'

    if request.method == 'POST':
        print("Login POST request received")
        username_or_email = request.form.get('username_or_email')
        password = request.form.get('password')
        print(f"Received username_or_email: {username_or_email}")

        if not username_or_email or not password:
            flash('Please fill out all fields.')
            return redirect(url_for('login'))

        # Load workbook and find user by username or email
        if not os.path.exists(excel_file):
            flash('No users registered yet.')
            return redirect(url_for('login'))

        wb = openpyxl.load_workbook(excel_file)
        ws = wb['Users']

        user = None
        for row in ws.iter_rows(min_row=2, values_only=True):
            uname, email, pwd = row
            if uname == username_or_email or email == username_or_email:
                user = {'username': uname, 'email': email, 'password': pwd}
                break
        wb.close()

        if user and user['password'] == password:
            session['username'] = user['username']
            flash('Login successful.')
            # Redirect to painting page after successful login
            return redirect(url_for('painting'))
        else:
            flash('Invalid credentials.')
            return redirect(url_for('login'))

    return render_template('login.html')

@app.route('/logout')
def logout():
    session.pop('username', None)
    flash('You have been logged out.')
    return redirect(url_for('login'))

@app.route('/painting')
def painting():
    if 'username' not in session:
        flash('Please log in to access the painting page.')
        return redirect(url_for('login'))
    return render_template('painting.html')

@app.route('/video_feed')
def video_feed():
    return Response(generate_frames(), mimetype='multipart/x-mixed-replace; boundary=frame')

def generate_frames():
    global current_color, xp, yp, canvas, camera_active
    cap = None
    
    while True:
        if camera_active and cap is None:
            cap = cv2.VideoCapture(0)
        elif not camera_active and cap is not None:
            cap.release()
            cap = None
            continue
            
        if not camera_active or cap is None:
            continue
            
        success, frame = cap.read()
        if not success:
            continue
            
        try:
            # Process frame with MediaPipe
            frame = cv2.flip(frame, 1)
            rgb_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
            results = hands.process(rgb_frame)
        
            if results.multi_hand_landmarks:
                for hand_landmarks in results.multi_hand_landmarks:
                    # Get all landmarks
                    landmarks = []
                    for id, lm in enumerate(hand_landmarks.landmark):
                        h, w, c = frame.shape
                        cx, cy = int(lm.x * w), int(lm.y * h)
                        landmarks.append([id, cx, cy])
                    
                    if len(landmarks) != 0:
                        # Get finger tips
                        x1, y1 = landmarks[8][1], landmarks[8][2]  # Index finger
                        x2, y2 = landmarks[12][1], landmarks[12][2]  # Middle finger
                        
                        # Check fingers up
                        fingers = []
                        if landmarks[8][2] < landmarks[6][2]:  # Index finger up
                            fingers.append(1)
                        else:
                            fingers.append(0)
                            
                        if landmarks[12][2] < landmarks[10][2]:  # Middle finger up
                            fingers.append(1)
                        else:
                            fingers.append(0)
                        
                        # Selection mode (two fingers up)
                        if fingers == [1, 1]:
                            xp, yp = 0, 0
                            # Calculate distance between fingers for thickness
                            thickness = int(np.sqrt((x2-x1)**2 + (y2-y1)**2) / 5)
                            thickness = max(5, min(thickness, 50))  # Limit between 5-50
                            socketio.emit('size_change', {'thickness': thickness})
                            
                            # Check if selecting color from header
                            if y1 < 100:
                                if 0 < x1 < 100: current_color = (0, 0, 255)  # Red in BGR (Blue=0, Green=0, Red=255)
                                elif 100 < x1 < 200: current_color = (0, 255, 0)  # Green in BGR (Blue=0, Green=255, Red=0)
                                elif 200 < x1 < 300: current_color = (255, 0, 0)  # Blue in BGR (Blue=255, Green=0, Red=0)
                                elif 300 < x1 < 400: current_color = (0, 255, 255)  # Yellow in BGR (Blue=0, Green=255, Red=255)
                                elif 400 < x1 < 500: current_color = (255, 0, 255)  # Purple in BGR (Blue=255, Green=0, Red=255)
                                elif 500 < x1 < 600: current_color = (0, 0, 0)  # Black (eraser)
                                elif 600 < x1 < 700: current_color = (255, 255, 255)  # White
                        
                        # Drawing mode (index finger up)
                        elif fingers == [1, 0]:
                            if xp == 0 and yp == 0:
                                xp, yp = x1, y1
                            
                            # Use default thickness based on mode
                            thickness = 50 if current_color == (0, 0, 0) else 10
                            
                            cv2.line(canvas, (xp, yp), (x1, y1), current_color, thickness)
                            
                            # Convert BGR to RGB for canvas
                            rgb_color = (current_color[2], current_color[1], current_color[0])
                            socketio.emit('draw', {
                                'x1': xp,
                                'y1': yp,
                                'x2': x1,
                                'y2': y1,
                                'color': ','.join(map(str, rgb_color)),
                                'thickness': thickness
                            })
                            
                            xp, yp = x1, y1
                    
                    mp_drawing.draw_landmarks(frame, hand_landmarks, mp_hands.HAND_CONNECTIONS)

            # Combine frame and canvas
            frame = cv2.add(frame, canvas)
            ret, buffer = cv2.imencode('.jpg', frame)
            frame = buffer.tobytes()
            yield (b'--frame\r\n'
                   b'Content-Type: image/jpeg\r\n\r\n' + frame + b'\r\n')
        except Exception as e:
            print(f"Error processing frame: {e}")
            continue

@socketio.on('change_color')
def handle_change_color(data):
    global current_color
    r, g, b = map(int, data['color'].split(','))
    current_color = (b, g, r)  # Convert RGB to BGR for OpenCV

@socketio.on('start_cam')
def handle_start_cam():
    global camera_active
    camera_active = True

@socketio.on('stop_cam')
def handle_stop_cam():
    global camera_active
    camera_active = False

# Ensure drawing_files directory exists
os.makedirs('drawing_files', exist_ok=True)

@app.route('/save_drawing', methods=['POST'])
def save_drawing():
    if 'file' not in request.files:
        return {'status': 'error', 'message': 'No file provided'}, 400
    
    file = request.files['file']
    if file.filename == '':
        return {'status': 'error', 'message': 'No file selected'}, 400
    
    # Generate timestamp filename if none provided
    if file.filename == 'drawing.png':
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'drawing_{timestamp}.png'
    else:
        filename = file.filename
    
    save_path = os.path.join('drawing_files', filename)
    file.save(save_path)
    return {'status': 'success', 'filename': filename}

@socketio.on('clear_canvas')
def handle_clear_canvas():
    global canvas, xp, yp
    canvas = np.zeros((480, 640, 3), np.uint8)
    xp, yp = 0, 0
    socketio.emit('canvas_cleared')

@app.route('/help')
def help_page():
    return render_template('help.html')

import openpyxl
from markupsafe import Markup

@app.route('/view_users')
def view_users():
    excel_file = 'users.xlsx'
    if not os.path.exists(excel_file):
        flash('User data file not found.')
        return redirect(url_for('index'))

    wb = openpyxl.load_workbook(excel_file)
    ws = wb['Users']

    # Read all user data rows
    users_data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        users_data.append(row)
    wb.close()

    # Create an HTML table to display user data
    table_html = '<table class="table table-bordered"><thead><tr><th>Username</th><th>Email</th><th>Password</th></tr></thead><tbody>'
    for user in users_data:
        table_html += f'<tr><td>{user[0]}</td><td>{user[1]}</td><td>{user[2]}</td></tr>'
    table_html += '</tbody></table>'

    return render_template('view_users.html', table=Markup(table_html))

if __name__ == '__main__':
    socketio.run(app, debug=True)

@app.route('/video_feed')
def video_feed():
    return Response(generate_frames(), mimetype='multipart/x-mixed-replace; boundary=frame')

def generate_frames():
    global current_color, xp, yp, canvas, camera_active
    cap = None
    
    while True:
        if camera_active and cap is None:
            cap = cv2.VideoCapture(0)
        elif not camera_active and cap is not None:
            cap.release()
            cap = None
            continue
            
        if not camera_active or cap is None:
            continue
            
        success, frame = cap.read()
        if not success:
            continue
            
        try:
            # Process frame with MediaPipe
            frame = cv2.flip(frame, 1)
            rgb_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
            results = hands.process(rgb_frame)
        
            if results.multi_hand_landmarks:
                for hand_landmarks in results.multi_hand_landmarks:
                    # Get all landmarks
                    landmarks = []
                    for id, lm in enumerate(hand_landmarks.landmark):
                        h, w, c = frame.shape
                        cx, cy = int(lm.x * w), int(lm.y * h)
                        landmarks.append([id, cx, cy])
                    
                    if len(landmarks) != 0:
                        # Get finger tips
                        x1, y1 = landmarks[8][1], landmarks[8][2]  # Index finger
                        x2, y2 = landmarks[12][1], landmarks[12][2]  # Middle finger
                        
                        # Check fingers up
                        fingers = []
                        if landmarks[8][2] < landmarks[6][2]:  # Index finger up
                            fingers.append(1)
                        else:
                            fingers.append(0)
                            
                        if landmarks[12][2] < landmarks[10][2]:  # Middle finger up
                            fingers.append(1)
                        else:
                            fingers.append(0)
                        
                        # Selection mode (two fingers up)
                        if fingers == [1, 1]:
                            xp, yp = 0, 0
                            # Calculate distance between fingers for thickness
                            thickness = int(np.sqrt((x2-x1)**2 + (y2-y1)**2) / 5)
                            thickness = max(5, min(thickness, 50))  # Limit between 5-50
                            socketio.emit('size_change', {'thickness': thickness})
                            
                            # Check if selecting color from header
                            if y1 < 100:
                                if 0 < x1 < 100: current_color = (0, 0, 255)  # Red in BGR (Blue=0, Green=0, Red=255)
                                elif 100 < x1 < 200: current_color = (0, 255, 0)  # Green in BGR (Blue=0, Green=255, Red=0)
                                elif 200 < x1 < 300: current_color = (255, 0, 0)  # Blue in BGR (Blue=255, Green=0, Red=0)
                                elif 300 < x1 < 400: current_color = (0, 255, 255)  # Yellow in BGR (Blue=0, Green=255, Red=255)
                                elif 400 < x1 < 500: current_color = (255, 0, 255)  # Purple in BGR (Blue=255, Green=0, Red=255)
                                elif 500 < x1 < 600: current_color = (0, 0, 0)  # Black (eraser)
                                elif 600 < x1 < 700: current_color = (255, 255, 255)  # White
                        
                        # Drawing mode (index finger up)
                        elif fingers == [1, 0]:
                            if xp == 0 and yp == 0:
                                xp, yp = x1, y1
                            
                            # Use default thickness based on mode
                            thickness = 50 if current_color == (0, 0, 0) else 10
                            
                            cv2.line(canvas, (xp, yp), (x1, y1), current_color, thickness)
                            
                            # Convert BGR to RGB for canvas
                            rgb_color = (current_color[2], current_color[1], current_color[0])
                            socketio.emit('draw', {
                                'x1': xp,
                                'y1': yp,
                                'x2': x1,
                                'y2': y1,
                                'color': ','.join(map(str, rgb_color)),
                                'thickness': thickness
                            })
                            
                            xp, yp = x1, y1
                    
                    mp_drawing.draw_landmarks(frame, hand_landmarks, mp_hands.HAND_CONNECTIONS)

            # Combine frame and canvas
            frame = cv2.add(frame, canvas)
            ret, buffer = cv2.imencode('.jpg', frame)
            frame = buffer.tobytes()
            yield (b'--frame\r\n'
                   b'Content-Type: image/jpeg\r\n\r\n' + frame + b'\r\n')
        except Exception as e:
            print(f"Error processing frame: {e}")
            continue

@socketio.on('change_color')
def handle_change_color(data):
    global current_color
    r, g, b = map(int, data['color'].split(','))
    current_color = (b, g, r)  # Convert RGB to BGR for OpenCV

@socketio.on('start_cam')
def handle_start_cam():
    global camera_active
    camera_active = True

@socketio.on('stop_cam')
def handle_stop_cam():
    global camera_active
    camera_active = False

import os
from datetime import datetime
from flask import request, jsonify, send_from_directory
import subprocess
import platform

# Ensure drawing_files directory exists
os.makedirs('drawing_files', exist_ok=True)

@app.route('/save_drawing', methods=['POST'])
def save_drawing():
    if 'file' not in request.files:
        return {'status': 'error', 'message': 'No file provided'}, 400
    
    file = request.files['file']
    if file.filename == '':
        return {'status': 'error', 'message': 'No file selected'}, 400
    
    # Generate timestamp filename if none provided
    if file.filename == 'drawing.png':
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'drawing_{timestamp}.png'
    else:
        filename = file.filename
    
    save_path = os.path.join('drawing_files', filename)
    file.save(save_path)
    return {'status': 'success', 'filename': filename}

@app.route('/list_drawings', methods=['GET'])
def list_drawings():
    files = []
    for filename in os.listdir('drawing_files'):
        if filename.lower().endswith(('.png', '.jpg', '.jpeg')):
            files.append(filename)
    files.sort(reverse=True)  # latest first
    return jsonify(files)

@app.route('/load_drawing/<filename>', methods=['GET'])
def load_drawing(filename):
    return send_from_directory('drawing_files', filename)

@app.route('/open_drawing_folder', methods=['GET'])
def open_drawing_folder():
    folder_path = os.path.abspath('drawing_files')
    system = platform.system()
    try:
        if system == 'Windows':
            os.startfile(folder_path)
        elif system == 'Darwin':  # macOS
            subprocess.Popen(['open', folder_path])
        else:  # Linux and others
            subprocess.Popen(['xdg-open', folder_path])
        return jsonify({'status': 'success', 'message': f'Opened folder: {folder_path}'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@socketio.on('clear_canvas')
def handle_clear_canvas():
    global canvas, xp, yp
    canvas = np.zeros((480, 640, 3), np.uint8)
    xp, yp = 0, 0
    socketio.emit('canvas_cleared')

if __name__ == '__main__':
    socketio.run(app, debug=True)
